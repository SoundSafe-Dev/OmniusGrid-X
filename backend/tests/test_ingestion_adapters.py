"""Focused contract and safety tests for the correlation ingestion adapters."""

from io import BytesIO
from zipfile import ZIP_DEFLATED, ZipFile

from app.services.ingestion_adapters import (
    IngestionLimits,
    connector_capabilities,
    create_default_registry,
    ingest_file,
    plan_connector_ingestion,
)


def _zip_bytes(entries):
    output = BytesIO()
    with ZipFile(output, "w", ZIP_DEFLATED) as archive:
        for name, value in entries.items():
            archive.writestr(name, value)
    return output.getvalue()


def test_ingest_tsv_returns_common_table_contract_and_lineage():
    result = ingest_file(
        b"asset\tfacility\tdate\tminutes\nMX-101\tPlant A\t2026-08-01\t12\n",
        "operations.tsv",
    )

    assert result["status"] == "parsed"
    assert result["errors"] == []
    assert result["tables"] == {
        "Sheet1": [{
            "asset": "MX-101",
            "facility": "Plant A",
            "date": "2026-08-01",
            "minutes": "12",
        }]
    }
    assert result["manifest"]["source"]["format"]["format"] == "tsv"
    assert result["manifest"]["source"]["sha256"]
    assert result["manifest"]["tables"][0]["source_table"] == "Sheet1"
    assert result["capabilities"]["selected_format"]["supports_tables"] is True


def test_json_and_jsonl_are_bounded_table_sources_without_false_mismatch_warning():
    json_result = ingest_file(
        b'{"work_orders":[{"asset":"MX-1","hours":2}],"facility":"A"}',
        "work_orders.json",
    )
    jsonl_result = ingest_file(
        b'{"asset":"MX-1","hours":2}\n{"asset":"MX-2","hours":3}\n',
        "work_orders.jsonl",
    )

    assert json_result["tables"]["work_orders"] == [{"asset": "MX-1", "hours": 2}]
    assert json_result["manifest"]["tables"][0]["source_metadata"]["root_metadata"] == {"facility": "A"}
    assert jsonl_result["status"] == "parsed"
    assert not any(warning["code"] == "filename_content_mismatch" for warning in jsonl_result["warnings"])
    assert len(jsonl_result["tables"]["records"]) == 2


def test_xml_rejects_entity_declarations_and_parses_direct_records():
    safe_result = ingest_file(
        b"<records><record id='a'><asset>MX-1</asset><hours>2</hours></record></records>",
        "records.xml",
    )
    unsafe_result = ingest_file(
        b"<!DOCTYPE data [<!ENTITY boom 'unsafe'>]><data><record>&boom;</record></data>",
        "unsafe.xml",
    )

    assert safe_result["tables"]["record"] == [{"@id": "a", "asset": "MX-1", "hours": "2"}]
    assert unsafe_result["status"] == "rejected"
    assert unsafe_result["errors"][0]["code"] == "xml_dtd_not_allowed"


def test_xlsm_is_read_as_a_multi_sheet_workbook_without_executing_macros():
    from openpyxl import Workbook

    workbook = Workbook()
    production = workbook.active
    production.title = "Production"
    production.append(["Asset", "Downtime"])
    production.append(["MX-1", 12])
    quality = workbook.create_sheet("Quality")
    quality.append(["Asset", "Defects"])
    quality.append(["MX-1", 2])
    payload = BytesIO()
    workbook.save(payload)

    result = ingest_file(payload.getvalue(), "operations.xlsm")

    assert result["status"] == "parsed"
    assert result["tables"]["Production"] == [{"Asset": "MX-1", "Downtime": 12}]
    assert result["tables"]["Quality"] == [{"Asset": "MX-1", "Defects": 2}]
    assert result["manifest"]["formula_mode"] == "cached_values_only"


def test_workbook_table_cap_is_explicit_without_parsing_later_sheets():
    from openpyxl import Workbook

    workbook = Workbook()
    first = workbook.active
    first.title = "Production"
    first.append(["Asset", "Downtime"])
    first.append(["MX-1", 12])
    second = workbook.create_sheet("Quality")
    second.append(["Asset", "Defects"])
    second.append(["MX-1", 2])
    payload = BytesIO()
    workbook.save(payload)

    result = ingest_file(
        payload.getvalue(),
        "operations.xlsx",
        limits=IngestionLimits(max_tables=1),
    )

    assert result["status"] == "partial"
    assert result["manifest"]["truncated"] is True
    assert result["manifest"]["table_limit"]["dropped_table_count"] == 1
    assert result["manifest"]["table_limit"]["unparsed_source_count"] == 0
    assert list(result["tables"]) == ["Production"]
    assert [warning["code"] for warning in result["warnings"]].count("table_limit_reached") == 1


def test_zip_batch_safely_extracts_structured_children_and_validates_zip_bounds():
    batch = _zip_bytes({
        "exports/operations.csv": "asset,value\nMX-1,1\n",
        "exports/quality.json": '{"defects": 2}',
    })
    result = ingest_file(batch, "batch.zip")
    bounded = ingest_file(batch, "batch.zip", limits=IngestionLimits(max_zip_entries=1))

    assert result["status"] == "parsed"
    assert result["tables"]["exports_operations.csv__Sheet1"] == [{"asset": "MX-1", "value": "1"}]
    assert result["tables"]["exports_quality.json__json"] == [{"defects": 2}]
    assert result["manifest"]["batch_manifest"]["entry_count"] == 2
    assert result["manifest"]["batch_manifest"]["extraction_performed"] is True
    assert bounded["errors"][0]["code"] == "zip_entry_limit_exceeded"


def test_zip_batch_allowlist_parses_only_selected_normalized_paths_without_using_capacity():
    batch = _zip_bytes({
        "exports/one.csv": "asset,value\nMX-1,1\n",
        "exports/two.csv": "asset,value\nMX-2,2\n",
        "exports/three.csv": "asset,value\nMX-3,3\n",
    })

    result = ingest_file(
        batch,
        "batch.zip",
        limits=IngestionLimits(max_tables=1, max_total_rows=1),
        # Canonicalization is deliberate: selection compares normalized
        # internal paths, never a client-provided raw ZIP member string.
        archive_entry_allowlist=["exports//two.csv"],
    )
    batch_manifest = result["manifest"]["batch_manifest"]
    one, two, three = batch_manifest["children"]

    assert result["status"] == "parsed"
    assert result["manifest"]["truncated"] is False
    assert result["tables"] == {
        "exports_two.csv__Sheet1": [{"asset": "MX-2", "value": "2"}]
    }
    assert batch_manifest["entries"][1]["normalized_path"] == "exports/two.csv"
    assert batch_manifest["selection"] == {
        "archive_entry_allowlist_applied": True,
        "selected_normalized_paths": ["exports/two.csv"],
        "selected_entry_count": 1,
        "not_selected_entry_count": 2,
    }
    retained_schema = result["manifest"]["tables"][0]
    assert retained_schema["source_metadata"]["archive_path"] == "exports/two.csv"
    assert retained_schema["source_metadata"]["normalized_archive_path"] == "exports/two.csv"
    assert (one["status"], one["reason"]) == (
        "not_selected", "archive_entry_not_selected"
    )
    assert two["status"] == "parsed"
    assert (three["status"], three["reason"]) == (
        "not_selected", "archive_entry_not_selected"
    )


def test_zip_batch_allowlist_rejects_unknown_or_unsafe_paths_and_non_zip_inputs():
    batch = _zip_bytes({"exports/one.csv": "asset,value\nMX-1,1\n"})

    unknown_path = ingest_file(
        batch,
        "batch.zip",
        archive_entry_allowlist=["exports/missing.csv"],
    )
    unsafe_path = ingest_file(
        batch,
        "batch.zip",
        archive_entry_allowlist=["../exports/one.csv"],
    )
    non_zip = ingest_file(
        b"asset,value\nMX-1,1\n",
        "operations.csv",
        archive_entry_allowlist=["exports/one.csv"],
    )

    assert unknown_path["errors"][0]["code"] == "archive_entry_not_found"
    assert unsafe_path["errors"][0]["code"] == "invalid_archive_entry_allowlist"
    assert non_zip["errors"][0]["code"] == "archive_entry_allowlist_requires_zip"


def test_zip_batch_table_cap_marks_manifest_partial_and_skips_later_children():
    batch = _zip_bytes({
        "exports/one.csv": "asset,value\nMX-1,1\n",
        "exports/two.csv": "asset,value\nMX-2,2\n",
        "exports/three.csv": "asset,value\nMX-3,3\n",
    })

    result = ingest_file(batch, "batch.zip", limits=IngestionLimits(max_tables=1))
    batch_manifest = result["manifest"]["batch_manifest"]
    table_limit = result["manifest"]["table_limit"]

    assert result["status"] == "partial"
    assert result["manifest"]["truncated"] is True
    assert table_limit == {
        "max_tables": 1,
        "retained_table_count": 1,
        "dropped_table_count": 0,
        "unparsed_source_count": 2,
        "capacity_reached": True,
        "truncated": True,
    }
    assert batch_manifest["truncated"] is True
    assert batch_manifest["table_limit"] == table_limit
    assert batch_manifest["children"][0]["status"] == "parsed"
    for child in batch_manifest["children"][1:]:
        assert child["status"] == "not_retained"
        assert child["reason"] == "table_limit_reached_before_parse"
        assert child["truncated"] is True
        assert child["tables"] == []
    assert [warning["code"] for warning in result["warnings"]].count("table_limit_reached") == 1


def test_zip_batch_reports_tables_dropped_inside_a_partially_retained_child():
    batch = _zip_bytes({
        "exports/multi.json": (
            '{"work_orders":[{"asset":"MX-1"}],'
            '"maintenance":[{"asset":"MX-1"}]}'
        ),
        "exports/later.csv": "asset,value\nMX-2,2\n",
    })

    result = ingest_file(batch, "batch.zip", limits=IngestionLimits(max_tables=1))
    batch_manifest = result["manifest"]["batch_manifest"]
    first_child, later_child = batch_manifest["children"]

    assert result["status"] == "partial"
    assert result["manifest"]["truncated"] is True
    assert result["manifest"]["table_limit"]["dropped_table_count"] == 1
    assert result["manifest"]["table_limit"]["unparsed_source_count"] == 1
    assert first_child["status"] == "partial"
    assert first_child["reason"] == "some_tables_not_retained_due_to_table_limit"
    assert first_child["dropped_table_count"] == 1
    assert len(first_child["tables"]) == 1
    assert later_child["status"] == "not_retained"
    assert later_child["reason"] == "table_limit_reached_before_parse"
    assert [warning["code"] for warning in result["warnings"]].count("table_limit_reached") == 1


def test_optional_and_conversion_formats_return_clear_capability_errors():
    parquet_result = ingest_file(b"not parquet", "metrics.parquet")
    numbers_result = ingest_file(_zip_bytes({"Index/Document.iwa": b"placeholder"}), "plant.numbers")

    assert parquet_result["errors"][0]["code"] in {
        "optional_dependency_missing",
        "parquet_parse_failed",
    }
    assert numbers_result["errors"][0]["code"] == "numbers_conversion_required"
    assert numbers_result["capabilities"]["selected_format"]["parse_mode"] == "conversion_required"


def test_opt_in_ocr_adapter_and_external_connector_plan_are_non_networked():
    class FakeOcr:
        name = "test-ocr"

        def extract(self, content, filename, *, limits):
            return {"tables": {"ocr_rows": [{"asset": "MX-1", "text": "alarm"}]}}

    registry = create_default_registry()
    registry.set_ocr_adapter(FakeOcr())
    result = ingest_file(b"\x89PNG\r\n\x1a\nimage", "alarm.png", registry=registry, enable_ocr=True)
    plan = plan_connector_ingestion(
        "postgres",
        {"host": "db.internal", "database": "ops", "credential_ref": "secret://db", "password": "not-returned"},
        entities=["work_orders"],
    )

    assert result["status"] == "parsed"
    assert result["tables"]["ocr_rows"] == [{"asset": "MX-1", "text": "alarm"}]
    assert plan["connection_attempted"] is False
    assert "password" in plan["provided_configuration_keys"]
    assert "not-returned" not in str(plan)
    assert {capability["category"] for capability in connector_capabilities()} >= {
        "database", "warehouse", "erp", "historian", "event_stream"
    }


def test_shared_row_cap_is_enforced_with_a_partial_result():
    result = ingest_file(
        b"asset,value\nMX-1,1\nMX-2,2\nMX-3,3\n",
        "small.csv",
        limits={"max_total_rows": 2, "max_rows_per_table": 10},
    )

    assert result["status"] == "partial"
    assert len(result["tables"]["Sheet1"]) == 2
    assert result["manifest"]["truncated"] is True
    assert any(warning["code"] == "total_row_limit_reached" for warning in result["warnings"])
