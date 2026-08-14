"""Safe, extensible ingestion adapters for operational correlation data.

This module deliberately keeps file reading deterministic.  It turns a local
upload into bounded, JSON-safe tables and a manifest that preserves enough
lineage for the common-evidence layer to reason about the source later.

It does *not* open network connections or execute macros, formulas, OCR, or
legacy-document conversions by default.  Those integrations are represented by
explicit interfaces so a deployment can opt into an approved connector/worker.

The public entry point is :func:`ingest_file`::

    result = ingest_file(content, "production.tsv")
    rows = result["tables"]["production"]

Every result includes ``tables``, ``manifest``, ``capabilities``, ``warnings``
and ``errors``.  Expected limitations are therefore inspectable instead of
being confused with an empty, successful parse.
"""

from __future__ import annotations

import csv
import hashlib
import importlib.util
import io
import json
import math
import re
import zipfile
from abc import ABC, abstractmethod
from collections import OrderedDict
from dataclasses import asdict, dataclass, field, replace
from datetime import date, datetime, time
from decimal import Decimal
from pathlib import PurePosixPath
from typing import Any, Dict, Iterable, Iterator, List, Mapping, Optional, Protocol, Sequence, Tuple, Union


# These defaults are intentionally conservative for request/worker use.  A
# batch worker may supply larger, explicitly approved limits; it should never
# need to fork a parser just to change a bound.
DEFAULT_MAX_FILE_BYTES = 50 * 1024 * 1024


@dataclass(frozen=True)
class IngestionLimits:
    """Hard safety bounds applied before and during parsing.

    ``max_total_rows`` is shared by all tables in one source.  It prevents a
    many-sheet workbook or a deeply nested JSON file from bypassing the
    per-table row cap.
    """

    max_file_bytes: int = DEFAULT_MAX_FILE_BYTES
    max_tables: int = 100
    max_rows_per_table: int = 100_000
    max_total_rows: int = 250_000
    max_columns: int = 1_000
    max_cell_chars: int = 100_000
    max_json_depth: int = 64
    max_json_line_chars: int = 1_000_000
    max_xml_depth: int = 64
    max_xml_elements: int = 1_000_000
    max_document_pages: int = 100
    max_zip_entries: int = 1_000
    max_zip_uncompressed_bytes: int = 512 * 1024 * 1024
    max_zip_compression_ratio: float = 200.0

    def __post_init__(self) -> None:
        for name, value in asdict(self).items():
            if isinstance(value, (int, float)) and value <= 0:
                raise ValueError(f"{name} must be greater than zero")

    def as_dict(self) -> Dict[str, Union[int, float]]:
        return asdict(self)


LimitsInput = Optional[Union[IngestionLimits, Mapping[str, Any]]]
ArchiveEntryAllowlistInput = Optional[Sequence[str]]


@dataclass(frozen=True)
class IngestionIssue:
    """A machine-readable warning/error with a user-safe explanation."""

    code: str
    message: str
    remediation: Optional[str] = None
    retryable: bool = False

    def as_dict(self) -> Dict[str, Any]:
        data: Dict[str, Any] = {"code": self.code, "message": self.message}
        if self.remediation:
            data["remediation"] = self.remediation
        if self.retryable:
            data["retryable"] = True
        return data


class IngestionFailure(Exception):
    """Expected input/capability failure that should become a structured result."""

    def __init__(
        self,
        code: str,
        message: str,
        remediation: Optional[str] = None,
        *,
        retryable: bool = False,
    ) -> None:
        super().__init__(message)
        self.issue = IngestionIssue(code, message, remediation, retryable)


@dataclass(frozen=True)
class FormatDescriptor:
    """A format classification based on a filename, declared type and bytes."""

    format: str
    family: str
    extension: str = ""
    confidence: str = "extension"
    declared_content_type: Optional[str] = None
    magic_format: Optional[str] = None
    mismatch: bool = False

    def as_dict(self) -> Dict[str, Any]:
        return {
            "format": self.format,
            "family": self.family,
            "extension": self.extension,
            "confidence": self.confidence,
            "declared_content_type": self.declared_content_type,
            "magic_format": self.magic_format,
            "mismatch": self.mismatch,
        }


@dataclass(frozen=True)
class FormatCapability:
    """A format's explicit parse contract, including optional dependencies."""

    format: str
    family: str
    extensions: Tuple[str, ...]
    parse_mode: str
    supports_tables: bool
    dependency: Tuple[str, ...] = ()
    available: bool = True
    notes: str = ""

    def as_dict(self) -> Dict[str, Any]:
        return {
            "format": self.format,
            "family": self.family,
            "extensions": list(self.extensions),
            "parse_mode": self.parse_mode,
            "supports_tables": self.supports_tables,
            "dependency": list(self.dependency),
            "available": self.available,
            "notes": self.notes,
        }


@dataclass
class ParseOutput:
    """Adapter-specific supplemental manifest fields."""

    manifest: Dict[str, Any] = field(default_factory=dict)


def _module_available(module: str) -> bool:
    return importlib.util.find_spec(module) is not None


def _coerce_limits(limits: LimitsInput) -> IngestionLimits:
    if limits is None:
        return IngestionLimits()
    if isinstance(limits, IngestionLimits):
        return limits
    if isinstance(limits, Mapping):
        allowed = set(IngestionLimits.__dataclass_fields__)
        unexpected = sorted(set(limits) - allowed)
        if unexpected:
            raise ValueError(f"unknown ingestion limit(s): {', '.join(unexpected)}")
        return IngestionLimits(**dict(limits))
    raise ValueError("limits must be an IngestionLimits instance or mapping")


def _normalise_archive_entry_path(value: Any) -> str:
    """Return one safe, canonical ZIP-member path without reading the entry."""

    if not isinstance(value, str):
        raise ValueError("archive entry paths must be strings")
    name = value.replace("\\", "/")
    path = PurePosixPath(name)
    normalised = str(path)
    if (
        not name
        or "\x00" in name
        or path.is_absolute()
        or ".." in path.parts
        or normalised in {"", "."}
    ):
        raise ValueError("archive entry paths must be non-empty relative paths without parent segments")
    return normalised


def _coerce_archive_entry_allowlist(
    allowlist: ArchiveEntryAllowlistInput,
    *,
    max_entries: int,
) -> Optional[frozenset[str]]:
    """Validate an optional, bounded set of ZIP member paths from a caller."""

    if allowlist is None:
        return None
    if isinstance(allowlist, (str, bytes, bytearray)):
        raise ValueError("archive_entry_allowlist must be a sequence of archive paths, not a string")
    try:
        raw_paths = iter(allowlist)
    except TypeError as exc:
        raise ValueError("archive_entry_allowlist must be a sequence of archive paths") from exc

    normalised_paths = set()
    for index, raw_path in enumerate(raw_paths, start=1):
        if index > max_entries:
            raise ValueError(
                f"archive_entry_allowlist has more than the {max_entries} permitted archive paths"
            )
        normalised_paths.add(_normalise_archive_entry_path(raw_path))
    return frozenset(normalised_paths)


def _issue(code: str, message: str, remediation: Optional[str] = None) -> Dict[str, Any]:
    return IngestionIssue(code, message, remediation).as_dict()


def _safe_text(value: Any, max_chars: int) -> Tuple[Any, bool]:
    """Return a JSON-safe scalar and whether a source value was truncated."""

    if value is None:
        return None, False
    if isinstance(value, (datetime, date, time)):
        return value.isoformat(), False
    if isinstance(value, Decimal):
        # Decimal is not JSON serializable.  Preserve exact precision instead
        # of silently converting financial/measurement values to a float.
        return str(value), False
    if isinstance(value, bool):
        return value, False
    if isinstance(value, int):
        return value, False
    if isinstance(value, float):
        return (value if math.isfinite(value) else None), not math.isfinite(value)
    if isinstance(value, bytes):
        value = f"<binary:{len(value)} bytes>"
    elif isinstance(value, (dict, list, tuple, set)):
        value = json.dumps(value, ensure_ascii=False, default=str, separators=(",", ":"))
    else:
        value = str(value)
    if len(value) > max_chars:
        return value[:max_chars], True
    return value, False


def _normalise_columns(columns: Sequence[Any], max_columns: int) -> Tuple[List[str], bool]:
    """Create non-empty, unique and stable JSON column names."""

    normalised: List[str] = []
    seen: Dict[str, int] = {}
    truncated = False
    for index, value in enumerate(columns):
        if index >= max_columns:
            truncated = True
            break
        base = str(value or "").strip() or f"column_{index + 1}"
        # Avoid arbitrarily huge headers becoming a storage/response problem.
        base = base[:512]
        count = seen.get(base, 0) + 1
        seen[base] = count
        normalised.append(base if count == 1 else f"{base}_{count}")
    return normalised, truncated


def _normalise_table_name(name: Any, existing: Mapping[str, Any]) -> str:
    base = str(name or "table").strip() or "table"
    base = re.sub(r"[\x00-\x1f]", " ", base)[:256]
    candidate = base
    suffix = 2
    while candidate in existing:
        candidate = f"{base}_{suffix}"
        suffix += 1
    return candidate


class TableCollector:
    """Centralizes table/row/cell bounds across every adapter."""

    def __init__(self, limits: IngestionLimits) -> None:
        self.limits = limits
        self.tables: "OrderedDict[str, List[Dict[str, Any]]]" = OrderedDict()
        self.schemas: "OrderedDict[str, Dict[str, Any]]" = OrderedDict()
        self.warnings: List[IngestionIssue] = []
        self.total_rows = 0
        # Keep table-cap accounting separate from ordinary parser warnings so
        # a large archive produces one useful summary rather than hundreds of
        # identical ``table_limit_reached`` messages.
        self.dropped_table_count = 0
        self.unparsed_source_count = 0
        self.external_truncated_source_count = 0

    @property
    def at_table_capacity(self) -> bool:
        return len(self.tables) >= self.limits.max_tables

    @property
    def table_limit_truncated(self) -> bool:
        return bool(self.dropped_table_count or self.unparsed_source_count)

    def note_tables_not_retained(self, count: int = 1) -> None:
        """Record a known number of tables omitted because of ``max_tables``.

        Adapters call this before breaking out of a multi-table parse.  It
        makes a table cap visible in the result manifest even when no rejected
        table ever reaches :meth:`add_table`.
        """

        self.dropped_table_count += max(0, int(count))

    def note_unparsed_source_due_to_table_limit(self, count: int = 1) -> None:
        """Record archive children intentionally left unread after the cap.

        These are sources whose exact table count is unknown because reading
        them would defeat the purpose of the capacity guard.  They still make
        the batch partial and are listed individually in its child manifest.
        """

        self.unparsed_source_count += max(0, int(count))

    def note_external_truncation(self, count: int = 1) -> None:
        """Record a child parser that was already truncated before merging."""

        self.external_truncated_source_count += max(0, int(count))

    def table_limit_manifest(self) -> Dict[str, Any]:
        """Return JSON-safe, explicit accounting for the shared table cap."""

        return {
            "max_tables": self.limits.max_tables,
            "retained_table_count": len(self.tables),
            "dropped_table_count": self.dropped_table_count,
            "unparsed_source_count": self.unparsed_source_count,
            "capacity_reached": self.at_table_capacity,
            "truncated": self.table_limit_truncated,
        }

    def all_warnings(self) -> List[IngestionIssue]:
        """Return ordinary warnings plus one deferred table-cap summary."""

        warnings = list(self.warnings)
        if self.table_limit_truncated:
            details = [
                f"retained {len(self.tables)} of at most {self.limits.max_tables} tables"
            ]
            if self.dropped_table_count:
                details.append(
                    f"did not retain {self.dropped_table_count} parsed table"
                    f"{'s' if self.dropped_table_count != 1 else ''}"
                )
            if self.unparsed_source_count:
                details.append(
                    f"did not parse {self.unparsed_source_count} archive entr"
                    f"{'ies' if self.unparsed_source_count != 1 else 'y'} after the table cap"
                )
            warnings.append(IngestionIssue(
                "table_limit_reached",
                "; ".join(details) + ".",
                "Split the workbook/batch, select the needed tables, or raise max_tables in an approved worker configuration.",
            ))
        if self.external_truncated_source_count:
            warnings.append(IngestionIssue(
                "child_source_truncated",
                f"{self.external_truncated_source_count} archived child source"
                f"{'s were' if self.external_truncated_source_count != 1 else ' was'} parsed with its own safety bounds.",
                "Inspect the batch child manifest before treating the batch as complete.",
            ))
        return warnings

    def add_table(
        self,
        name: Any,
        columns: Sequence[Any],
        rows: Iterable[Union[Mapping[str, Any], Sequence[Any]]],
        *,
        source_table: Optional[str] = None,
        source_metadata: Optional[Mapping[str, Any]] = None,
    ) -> str:
        if self.at_table_capacity:
            self.note_tables_not_retained()
            return ""

        table_name = _normalise_table_name(name, self.tables)
        clean_columns, columns_truncated = _normalise_columns(columns, self.limits.max_columns)
        if columns_truncated:
            self.warnings.append(IngestionIssue(
                "column_limit_reached",
                f"Table '{table_name}' has more than {self.limits.max_columns} columns; extra columns were omitted.",
            ))

        accepted: List[Dict[str, Any]] = []
        truncated = columns_truncated
        cell_truncated = False
        row_limit_warning_emitted = False
        for raw_row in rows:
            if len(accepted) >= self.limits.max_rows_per_table:
                truncated = True
                if not row_limit_warning_emitted:
                    self.warnings.append(IngestionIssue(
                        "row_limit_reached",
                        f"Table '{table_name}' was capped at {self.limits.max_rows_per_table} rows.",
                        "Use an asynchronous batch job or a higher approved limit for the full source.",
                    ))
                    row_limit_warning_emitted = True
                break
            if self.total_rows >= self.limits.max_total_rows:
                truncated = True
                self.warnings.append(IngestionIssue(
                    "total_row_limit_reached",
                    f"The source was capped at {self.limits.max_total_rows} total rows across tables.",
                    "Use an asynchronous batch job or a higher approved limit for the full source.",
                ))
                break

            if isinstance(raw_row, Mapping):
                values = [raw_row.get(column) for column in clean_columns]
                # JSON rows may have original keys before they were normalized.
                if not any(value is not None for value in values):
                    values = [raw_row.get(str(column)) for column in clean_columns]
            else:
                values = list(raw_row)
            record: Dict[str, Any] = {}
            for index, column in enumerate(clean_columns):
                value = values[index] if index < len(values) else None
                safe_value, was_truncated = _safe_text(value, self.limits.max_cell_chars)
                record[column] = safe_value
                cell_truncated = cell_truncated or was_truncated
            accepted.append(record)
            self.total_rows += 1

        if cell_truncated:
            self.warnings.append(IngestionIssue(
                "cell_value_truncated",
                f"One or more values in '{table_name}' exceeded {self.limits.max_cell_chars} characters.",
            ))

        self.tables[table_name] = accepted
        self.schemas[table_name] = {
            "name": table_name,
            "source_table": str(source_table or name or table_name),
            "columns": clean_columns,
            "row_count": len(accepted),
            "truncated": truncated,
            "source_metadata": dict(source_metadata or {}),
        }
        return table_name


@dataclass
class ParseContext:
    source: FormatDescriptor
    filename: str
    limits: IngestionLimits
    collector: TableCollector
    registry: "IngestionAdapterRegistry"
    enable_ocr: bool = False
    enable_legacy_conversion: bool = False
    # This is intentionally meaningful only to the top-level ZIP batch
    # adapter.  Recursive child parsing never inherits it.
    archive_entry_allowlist: Optional[frozenset[str]] = None
    warnings: List[IngestionIssue] = field(default_factory=list)


class OCRAdapter(Protocol):
    """Approved OCR implementation interface.

    Implementations may call a local binary or an approved service, but they
    must be registered deliberately.  They receive bounded input and return a
    JSON-safe mapping, ideally with ``text`` and/or ``tables``.
    """

    name: str

    def extract(
        self, content: bytes, filename: str, *, limits: IngestionLimits
    ) -> Mapping[str, Any]:
        ...


class LegacyDocConverter(Protocol):
    """Conversion worker interface for legacy binary ``.doc`` files."""

    name: str

    def convert(
        self, content: bytes, filename: str, *, limits: IngestionLimits
    ) -> Mapping[str, Any]:
        ...


class IngestionAdapter(ABC):
    """Base class for deterministic local format adapters."""

    formats: Tuple[str, ...] = ()

    @abstractmethod
    def capability(self, selected_format: str) -> FormatCapability:
        """Describe the selected format without parsing it."""

    @abstractmethod
    def parse(self, content: bytes, context: ParseContext) -> ParseOutput:
        """Populate ``context.collector`` or raise :class:`IngestionFailure`."""


class IngestionAdapterRegistry:
    """A pluggable adapter registry; no custom adapter is run implicitly."""

    def __init__(self) -> None:
        self._adapters: Dict[str, IngestionAdapter] = {}
        self.ocr_adapter: Optional[OCRAdapter] = None
        self.legacy_doc_converter: Optional[LegacyDocConverter] = None

    def register(self, adapter: IngestionAdapter) -> None:
        for format_name in adapter.formats:
            if format_name in self._adapters:
                raise ValueError(f"an adapter is already registered for '{format_name}'")
            self._adapters[format_name] = adapter

    def get(self, format_name: str) -> Optional[IngestionAdapter]:
        return self._adapters.get(format_name)

    def capabilities(self) -> List[Dict[str, Any]]:
        return [
            self._adapters[name].capability(name).as_dict()
            for name in sorted(self._adapters)
        ]

    def set_ocr_adapter(self, adapter: Optional[OCRAdapter]) -> None:
        self.ocr_adapter = adapter

    def set_legacy_doc_converter(self, converter: Optional[LegacyDocConverter]) -> None:
        self.legacy_doc_converter = converter


_EXTENSION_FORMATS: Dict[str, Tuple[str, str]] = {
    ".csv": ("csv", "tabular"),
    ".tsv": ("tsv", "tabular"),
    ".tab": ("tsv", "tabular"),
    ".txt": ("delimited", "tabular"),
    ".json": ("json", "structured"),
    ".jsonl": ("jsonl", "structured"),
    ".ndjson": ("jsonl", "structured"),
    ".parquet": ("parquet", "columnar"),
    ".pq": ("parquet", "columnar"),
    ".arrow": ("arrow", "columnar"),
    ".feather": ("arrow", "columnar"),
    ".ipc": ("arrow", "columnar"),
    ".xml": ("xml", "structured"),
    ".xlsx": ("xlsx", "spreadsheet"),
    ".xlsm": ("xlsm", "spreadsheet"),
    ".xls": ("xls", "spreadsheet"),
    ".xlsb": ("xlsb", "spreadsheet"),
    ".ods": ("ods", "spreadsheet"),
    ".numbers": ("numbers", "spreadsheet"),
    ".zip": ("zip", "archive"),
    ".docx": ("docx", "document"),
    ".doc": ("doc", "document"),
    ".pdf": ("pdf", "document"),
    ".png": ("image", "image"),
    ".jpg": ("image", "image"),
    ".jpeg": ("image", "image"),
    ".tif": ("image", "image"),
    ".tiff": ("image", "image"),
    ".bmp": ("image", "image"),
    ".webp": ("image", "image"),
}

_CONTENT_TYPE_FORMATS: Dict[str, Tuple[str, str]] = {
    "text/csv": ("csv", "tabular"),
    "text/tab-separated-values": ("tsv", "tabular"),
    "application/json": ("json", "structured"),
    "application/x-ndjson": ("jsonl", "structured"),
    "application/ndjson": ("jsonl", "structured"),
    "application/parquet": ("parquet", "columnar"),
    "application/vnd.apache.parquet": ("parquet", "columnar"),
    "application/xml": ("xml", "structured"),
    "text/xml": ("xml", "structured"),
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet": ("xlsx", "spreadsheet"),
    "application/vnd.ms-excel.sheet.macroenabled.12": ("xlsm", "spreadsheet"),
    "application/vnd.ms-excel.sheet.binary.macroenabled.12": ("xlsb", "spreadsheet"),
    "application/vnd.oasis.opendocument.spreadsheet": ("ods", "spreadsheet"),
    "application/zip": ("zip", "archive"),
}


def _magic_format(content: bytes) -> Optional[Tuple[str, str]]:
    head = content[:65_536]
    if head.startswith(b"PAR1") and content[-4:] == b"PAR1":
        return "parquet", "columnar"
    if head.startswith(b"ARROW1") or content[-6:] == b"ARROW1":
        return "arrow", "columnar"
    if head.startswith(b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"):
        return "xls", "spreadsheet"
    if head.startswith(b"%PDF-"):
        return "pdf", "document"
    if head.startswith(b"\x89PNG\r\n\x1a\n") or head.startswith(b"\xff\xd8\xff"):
        return "image", "image"
    if head.startswith((b"II*\x00", b"MM\x00*")):
        return "image", "image"
    if head.startswith(b"PK\x03\x04") or head.startswith(b"PK\x05\x06"):
        return "zip", "archive"
    # Text signatures are only used after an explicit binary check.
    stripped = head.lstrip(b"\xef\xbb\xbf\xff\xfe\xfe\xff \t\r\n")
    if stripped.startswith((b"<?xml", b"<")):
        return "xml", "structured"
    if stripped.startswith((b"{", b"[")):
        return "json", "structured"
    return None


def detect_format(
    content: bytes,
    filename: str = "",
    content_type: Optional[str] = None,
) -> FormatDescriptor:
    """Detect a supported format without attempting a full parse.

    For ZIP-container formats (XLSX, XLSM, ODS, Numbers), the filename is
    intentionally more specific than generic ZIP magic.  A generic ``.zip``
    never gets unpacked by detection.
    """

    extension = ""
    filename = filename or ""
    if "." in filename.rsplit("/", 1)[-1]:
        extension = "." + filename.rsplit(".", 1)[-1].lower()
    declared = (content_type or "").split(";", 1)[0].strip().lower() or None
    by_extension = _EXTENSION_FORMATS.get(extension)
    by_declared = _CONTENT_TYPE_FORMATS.get(declared or "")
    by_magic = _magic_format(content)

    selected = by_extension or by_declared or by_magic or ("unknown", "unknown")
    confidence = "extension" if by_extension else ("content_type" if by_declared else ("magic" if by_magic else "unknown"))
    selected_format, family = selected
    magic_name = by_magic[0] if by_magic else None

    # A generic ZIP signature is expected for OpenXML/ODS/Numbers packages.
    compatible_zip_container = selected_format in {"xlsx", "xlsm", "xlsb", "ods", "numbers", "docx"}
    compatible_textual_signature = selected_format in {"json", "jsonl"} and magic_name == "json"
    mismatch = bool(
        by_magic
        and magic_name != selected_format
        and not (magic_name == "zip" and compatible_zip_container)
        and not compatible_textual_signature
    )
    return FormatDescriptor(
        format=selected_format,
        family=family,
        extension=extension,
        confidence=confidence,
        declared_content_type=declared,
        magic_format=magic_name,
        mismatch=mismatch,
    )


def _decode_text(content: bytes, context: ParseContext) -> str:
    """Decode common operational exports without silently claiming fidelity."""

    if content.startswith((b"\xff\xfe", b"\xfe\xff")):
        try:
            return content.decode("utf-16")
        except UnicodeDecodeError as exc:
            raise IngestionFailure("text_decode_failed", "The UTF-16 text file could not be decoded.") from exc
    try:
        return content.decode("utf-8-sig")
    except UnicodeDecodeError:
        context.warnings.append(IngestionIssue(
            "text_decoded_with_replacement",
            "The file was not valid UTF-8; undecodable characters were replaced.",
            "Export the source as UTF-8 for exact text fidelity.",
        ))
        return content.decode("utf-8", errors="replace")


def _first_nonempty_row(rows: Iterator[Sequence[Any]], scan_limit: int = 30) -> Tuple[List[Any], Iterator[Sequence[Any]]]:
    """Find a useful header row while accepting leading blank spreadsheet rows."""

    buffered: List[Sequence[Any]] = []
    header: List[Any] = []
    for _ in range(scan_limit):
        try:
            candidate = next(rows)
        except StopIteration:
            break
        values = list(candidate)
        if any(value not in (None, "") for value in values):
            header = values
            break
        buffered.append(values)
    # Blank leading rows have no data semantics, so intentionally do not replay
    # them.  The returned iterator starts immediately after the header.
    return header, rows


class DelimitedAdapter(IngestionAdapter):
    formats = ("csv", "tsv", "delimited")

    def capability(self, selected_format: str) -> FormatCapability:
        return FormatCapability(
            format=selected_format,
            family="tabular",
            extensions=(".csv", ".tsv", ".tab", ".txt"),
            parse_mode="native_bounded",
            supports_tables=True,
            notes="CSV dialect is detected from a bounded sample; TSV always uses a tab delimiter.",
        )

    def parse(self, content: bytes, context: ParseContext) -> ParseOutput:
        text = _decode_text(content, context)
        delimiter = "\t" if context.source.format == "tsv" else ","
        if context.source.format != "tsv":
            sample = text[:32_768]
            try:
                dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
                delimiter = dialect.delimiter
            except csv.Error:
                # Comma remains the predictable default for a one-column CSV
                # as well as a malformed dialect declaration.
                pass

        try:
            reader = csv.reader(io.StringIO(text), delimiter=delimiter)
            header, data_rows = _first_nonempty_row(iter(reader))
            if not header:
                context.collector.add_table("Sheet1", [], [])
                context.warnings.append(IngestionIssue(
                    "empty_tabular_file", "The delimited file contains no non-empty rows."
                ))
                return ParseOutput({"delimiter": delimiter})
            context.collector.add_table(
                "Sheet1",
                header,
                data_rows,
                source_table="Sheet1",
                source_metadata={"delimiter": delimiter},
            )
        except csv.Error as exc:
            raise IngestionFailure(
                "delimited_parse_failed",
                "The delimited file has an invalid CSV/TSV structure.",
                "Check quoting and delimiter settings, then export the source again.",
            ) from exc
        return ParseOutput({"delimiter": delimiter})


def _ensure_json_depth(value: Any, max_depth: int, depth: int = 0) -> None:
    if depth > max_depth:
        raise IngestionFailure(
            "json_depth_exceeded",
            f"JSON nesting exceeds the configured maximum depth of {max_depth}.",
            "Flatten the source or use a batch transformation before ingestion.",
        )
    if isinstance(value, Mapping):
        for child in value.values():
            _ensure_json_depth(child, max_depth, depth + 1)
    elif isinstance(value, (list, tuple)):
        for child in value:
            _ensure_json_depth(child, max_depth, depth + 1)


def _json_rows(value: Any, limit: Optional[int] = None) -> Iterator[Dict[str, Any]]:
    """Yield JSON records without duplicating an unbounded source array."""

    if isinstance(value, list):
        for index, item in enumerate(value):
            if limit is not None and index >= limit:
                break
            yield dict(item) if isinstance(item, Mapping) else {"value": item}
        return
    if isinstance(value, Mapping):
        yield dict(value)
        return
    yield {"value": value}


def _bounded_json_rows(value: Any, context: ParseContext) -> List[Dict[str, Any]]:
    # Keep one row beyond the collector's per-table cap so it can truthfully
    # report truncation without retaining the entire decoded JSON array again.
    limit = min(context.limits.max_rows_per_table, context.limits.max_total_rows) + 1
    return list(_json_rows(value, limit=limit))


def _columns_from_mapping_rows(rows: Sequence[Mapping[str, Any]], max_columns: int) -> Tuple[List[str], bool]:
    columns: List[str] = []
    seen = set()
    truncated = False
    for row in rows:
        for key in row:
            key_text = str(key)
            if key_text in seen:
                continue
            if len(columns) >= max_columns:
                truncated = True
                return columns, truncated
            seen.add(key_text)
            columns.append(key_text)
    return columns, truncated


class JsonAdapter(IngestionAdapter):
    formats = ("json", "jsonl")

    def capability(self, selected_format: str) -> FormatCapability:
        return FormatCapability(
            format=selected_format,
            family="structured",
            extensions=(".json", ".jsonl", ".ndjson"),
            parse_mode="native_bounded",
            supports_tables=True,
            notes="Top-level JSON arrays and arrays within an object become tables; nested values remain JSON-safe cell values.",
        )

    def _parse_jsonl(self, text: str, context: ParseContext) -> ParseOutput:
        rows: List[Dict[str, Any]] = []
        for line_number, line in enumerate(io.StringIO(text), start=1):
            if not line.strip():
                continue
            if len(line) > context.limits.max_json_line_chars:
                raise IngestionFailure(
                    "jsonl_line_too_large",
                    f"JSONL line {line_number} exceeds {context.limits.max_json_line_chars} characters.",
                    "Split oversized events before ingesting them.",
                )
            try:
                parsed = json.loads(line)
            except (json.JSONDecodeError, RecursionError) as exc:
                raise IngestionFailure(
                    "jsonl_parse_failed",
                    f"JSONL line {line_number} is not valid JSON.",
                    "Ensure each non-empty line is a complete JSON object/value.",
                ) from exc
            _ensure_json_depth(parsed, context.limits.max_json_depth)
            rows.extend(_json_rows(parsed, limit=1))
            if len(rows) > min(context.limits.max_rows_per_table, context.limits.max_total_rows):
                # Collector will expose the truncation uniformly; avoiding an
                # unnecessary parse of a very large JSONL request saves memory.
                break
        columns, truncated_columns = _columns_from_mapping_rows(rows, context.limits.max_columns)
        if truncated_columns:
            context.warnings.append(IngestionIssue(
                "column_limit_reached",
                f"JSONL contains more than {context.limits.max_columns} distinct fields; extra fields were omitted.",
            ))
        context.collector.add_table("records", columns, rows, source_table="records")
        return ParseOutput({"record_encoding": "jsonl"})

    def parse(self, content: bytes, context: ParseContext) -> ParseOutput:
        text = _decode_text(content, context)
        if context.source.format == "jsonl":
            return self._parse_jsonl(text, context)
        try:
            parsed = json.loads(text)
        except (json.JSONDecodeError, RecursionError) as exc:
            raise IngestionFailure(
                "json_parse_failed",
                "The file is not valid JSON.",
                "Validate the JSON export and upload it again.",
            ) from exc
        _ensure_json_depth(parsed, context.limits.max_json_depth)

        table_count = 0
        if isinstance(parsed, Mapping):
            # Operational JSON exports commonly wrap named record lists such as
            # {"work_orders": [...], "assets": [...]}.  Treat each one as a
            # source table and retain scalar root fields as metadata.
            root_scalars = {
                str(key): value
                for key, value in parsed.items()
                if not isinstance(value, (Mapping, list, tuple))
            }
            table_values = [
                (name, value)
                for name, value in parsed.items()
                if isinstance(value, list)
            ]
            for index, (name, value) in enumerate(table_values):
                if context.collector.at_table_capacity:
                    # Each selected top-level array is one deterministic
                    # table, so its omitted count is known without building
                    # any more bounded row buffers.
                    context.collector.note_tables_not_retained(len(table_values) - index)
                    break
                rows = _bounded_json_rows(value, context)
                columns, _ = _columns_from_mapping_rows(rows, context.limits.max_columns)
                context.collector.add_table(
                    name,
                    columns,
                    rows,
                    source_table=str(name),
                    source_metadata={"root_metadata": root_scalars},
                )
                table_count += 1
            if table_count == 0:
                rows = _bounded_json_rows(parsed, context)
                columns, _ = _columns_from_mapping_rows(rows, context.limits.max_columns)
                context.collector.add_table("json", columns, rows, source_table="json")
        else:
            rows = _bounded_json_rows(parsed, context)
            columns, _ = _columns_from_mapping_rows(rows, context.limits.max_columns)
            context.collector.add_table("records", columns, rows, source_table="records")

        return ParseOutput({"record_encoding": "json"})


def _xml_tag_name(tag: Any) -> str:
    text = str(tag)
    return text.rsplit("}", 1)[-1] if "}" in text else text


def _xml_record_to_row(element: Any, max_depth: int, depth: int = 0, prefix: str = "") -> Dict[str, Any]:
    if depth > max_depth:
        raise IngestionFailure(
            "xml_depth_exceeded",
            f"XML nesting exceeds the configured maximum depth of {max_depth}.",
            "Flatten the XML export before ingestion.",
        )
    row: Dict[str, Any] = {}
    for attr_name, attr_value in getattr(element, "attrib", {}).items():
        row[f"{prefix}@{_xml_tag_name(attr_name)}"] = attr_value
    children = list(element)
    if not children:
        key = prefix.rstrip(".") or _xml_tag_name(element.tag)
        row[key] = (element.text or "").strip()
        return row
    seen: Dict[str, int] = {}
    for child in children:
        name = _xml_tag_name(child.tag)
        seen[name] = seen.get(name, 0) + 1
        child_prefix = f"{prefix}{name}"
        if seen[name] > 1:
            child_prefix = f"{child_prefix}_{seen[name]}"
        child_prefix += "."
        row.update(_xml_record_to_row(child, max_depth, depth + 1, child_prefix))
    text = (element.text or "").strip()
    if text:
        row[f"{prefix.rstrip('.') or _xml_tag_name(element.tag)}_text"] = text
    return row


class XmlAdapter(IngestionAdapter):
    formats = ("xml",)

    def capability(self, selected_format: str) -> FormatCapability:
        return FormatCapability(
            format="xml",
            family="structured",
            extensions=(".xml",),
            parse_mode="native_streaming_bounded",
            supports_tables=True,
            dependency=("defusedxml (optional hardening)",),
            available=True,
            notes="DTD/entity declarations are rejected; direct children of the XML root become records.",
        )

    def parse(self, content: bytes, context: ParseContext) -> ParseOutput:
        # Reject declarations rather than allowing entity expansion in a parser
        # fallback.  This is checked across the bounded upload, not merely its
        # first few bytes, to avoid a padded-DOCTYPE bypass.
        upper_content = content.upper()
        if b"<!DOCTYPE" in upper_content or b"<!ENTITY" in upper_content:
            raise IngestionFailure(
                "xml_dtd_not_allowed",
                "XML files containing DTD or entity declarations are not accepted.",
                "Export XML without DTD/entity declarations.",
            )
        try:
            try:
                from defusedxml import ElementTree as safe_et  # type: ignore
            except ImportError:  # pragma: no cover - fallback remains protected above
                import xml.etree.ElementTree as safe_et  # type: ignore

            records: List[Dict[str, Any]] = []
            root_name = "xml"
            record_names: List[str] = []
            depth = 0
            element_count = 0
            # Iterparse lets us release each top-level record immediately.
            for event, element in safe_et.iterparse(io.BytesIO(content), events=("start", "end")):
                if event == "start":
                    depth += 1
                    element_count += 1
                    if element_count > context.limits.max_xml_elements:
                        raise IngestionFailure(
                            "xml_element_limit_reached",
                            f"XML exceeds {context.limits.max_xml_elements} elements.",
                            "Use a batch transformation or increase the approved XML element limit.",
                        )
                    if depth > context.limits.max_xml_depth:
                        raise IngestionFailure(
                            "xml_depth_exceeded",
                            f"XML nesting exceeds {context.limits.max_xml_depth} levels.",
                            "Flatten the XML export before ingestion.",
                        )
                    if depth == 1:
                        root_name = _xml_tag_name(element.tag) or "xml"
                else:
                    if depth == 2:
                        records.append(_xml_record_to_row(element, context.limits.max_xml_depth))
                        record_names.append(_xml_tag_name(element.tag))
                        element.clear()
                        if len(records) > context.limits.max_total_rows:
                            break
                    depth -= 1
            if not records:
                context.collector.add_table(root_name, [], [])
                context.warnings.append(IngestionIssue(
                    "xml_no_records", "No direct XML child elements were available as records."
                ))
                return ParseOutput({"root_element": root_name})
            columns, _ = _columns_from_mapping_rows(records, context.limits.max_columns)
            common_name = record_names[0] if len(set(record_names)) == 1 else root_name
            context.collector.add_table(
                common_name or "xml_records",
                columns,
                records,
                source_table=common_name or "xml_records",
                source_metadata={"root_element": root_name},
            )
            return ParseOutput({"root_element": root_name, "record_element": common_name})
        except IngestionFailure:
            raise
        except Exception as exc:
            raise IngestionFailure(
                "xml_parse_failed",
                "The file is not well-formed XML.",
                "Validate the XML export and upload it again.",
            ) from exc


def _safe_zip_entries(content: bytes, limits: IngestionLimits) -> List[zipfile.ZipInfo]:
    """Inspect a ZIP container without extracting any entry to disk."""

    try:
        with zipfile.ZipFile(io.BytesIO(content)) as archive:
            entries = archive.infolist()
    except (zipfile.BadZipFile, OSError) as exc:
        raise IngestionFailure(
            "invalid_zip_container",
            "The file is not a valid ZIP-based container.",
            "Re-export the source file or verify its extension.",
        ) from exc
    if len(entries) > limits.max_zip_entries:
        raise IngestionFailure(
            "zip_entry_limit_exceeded",
            f"The archive has {len(entries)} entries, above the limit of {limits.max_zip_entries}.",
            "Split the archive or increase the approved archive-entry limit.",
        )

    total_uncompressed = 0
    for entry in entries:
        try:
            _normalise_archive_entry_path(entry.filename)
        except ValueError as exc:
            raise IngestionFailure(
                "unsafe_zip_entry_path",
                "The archive contains an unsafe entry path.",
                "Create a ZIP without absolute or parent-directory paths.",
            ) from exc
        if entry.flag_bits & 0x1:
            raise IngestionFailure(
                "encrypted_zip_not_supported",
                "Encrypted ZIP containers cannot be inspected safely by this ingestion worker.",
                "Decrypt the archive in an approved workflow before uploading it.",
            )
        total_uncompressed += max(entry.file_size, 0)
        if total_uncompressed > limits.max_zip_uncompressed_bytes:
            raise IngestionFailure(
                "zip_uncompressed_size_exceeded",
                f"The archive expands beyond the {limits.max_zip_uncompressed_bytes} byte safety limit.",
                "Split the batch into smaller archives.",
            )
        if entry.file_size and not entry.is_dir():
            compressed = max(entry.compress_size, 1)
            ratio = entry.file_size / compressed
            if ratio > limits.max_zip_compression_ratio:
                raise IngestionFailure(
                    "zip_compression_ratio_exceeded",
                    "The archive has an unsafe compression ratio and may be a ZIP bomb.",
                    "Recreate the archive with normal compression or split it into smaller batches.",
                )
    return entries


def _zip_manifest_entry(info: zipfile.ZipInfo) -> Dict[str, Any]:
    descriptor = detect_format(b"", info.filename)
    return {
        "path": info.filename,
        "normalized_path": _normalise_archive_entry_path(info.filename),
        "is_directory": info.is_dir(),
        "compressed_bytes": info.compress_size,
        "uncompressed_bytes": info.file_size,
        "crc": f"{info.CRC:08x}",
        "format": descriptor.format,
        "family": descriptor.family,
        "nested_archive": descriptor.format == "zip",
    }


def _validate_archive_entry_allowlist(
    entries: Sequence[zipfile.ZipInfo],
    allowlist: Optional[frozenset[str]],
) -> None:
    """Require selected canonical paths to match one unambiguous ZIP member."""

    if allowlist is None:
        return
    available_paths = [_normalise_archive_entry_path(entry.filename) for entry in entries]
    available_path_set = set(available_paths)
    missing_paths = sorted(allowlist - available_path_set)
    if missing_paths:
        raise IngestionFailure(
            "archive_entry_not_found",
            "One or more selected archive paths are not present in this ZIP manifest.",
            "Refresh the archive manifest and select exact normalized entry paths.",
        )
    duplicate_paths = {
        path for path in available_path_set if available_paths.count(path) > 1
    }
    ambiguous_paths = sorted(allowlist & duplicate_paths)
    if ambiguous_paths:
        raise IngestionFailure(
            "archive_entry_path_ambiguous",
            "A selected archive path appears more than once after path normalization.",
            "Rebuild the ZIP without duplicate normalized paths before selecting entries.",
        )


class ZipBatchAdapter(IngestionAdapter):
    formats = ("zip",)

    def capability(self, selected_format: str) -> FormatCapability:
        return FormatCapability(
            format="zip",
            family="archive",
            extensions=(".zip",),
            parse_mode="safe_bounded_batch_extraction",
            supports_tables=True,
            notes=(
                "Inspects entries for path, encryption, compression-ratio, size, row, and table bounds; "
                "nested archives and conversion/OCR-only entries remain manifest-only."
            ),
        )

    def parse(self, content: bytes, context: ParseContext) -> ParseOutput:
        entries = _safe_zip_entries(content, context.limits)
        file_entries = [entry for entry in entries if not entry.is_dir()]
        _validate_archive_entry_allowlist(
            file_entries, context.archive_entry_allowlist
        )
        manifest_entries = [_zip_manifest_entry(entry) for entry in file_entries]
        total_uncompressed = sum(entry.file_size for entry in file_entries)
        children: List[Dict[str, Any]] = []

        # A batch should be convenient without becoming a recursive archive
        # interpreter.  Each child goes through the same detector and bounds as
        # a normal upload; nested archives, OCR, and conversion-only files are
        # retained in the manifest for an explicit worker workflow instead.
        with zipfile.ZipFile(io.BytesIO(content)) as archive:
            for entry in file_entries:
                normalized_path = _normalise_archive_entry_path(entry.filename)
                descriptor = detect_format(b"", entry.filename)
                child_summary: Dict[str, Any] = {
                    "path": entry.filename,
                    "normalized_path": normalized_path,
                    "format": descriptor.format,
                    "uncompressed_bytes": entry.file_size,
                    "status": "skipped",
                    "tables": [],
                }
                if (
                    context.archive_entry_allowlist is not None
                    and normalized_path not in context.archive_entry_allowlist
                ):
                    # Selection is intentional, so this neither consumes a
                    # table/row budget nor marks the selected batch partial.
                    child_summary["status"] = "not_selected"
                    child_summary["reason"] = "archive_entry_not_selected"
                    children.append(child_summary)
                    continue
                if descriptor.format == "zip":
                    child_summary["reason"] = "nested_archive_requires_explicit_review"
                    children.append(child_summary)
                    continue
                if entry.file_size > context.limits.max_file_bytes:
                    child_summary["reason"] = "child_file_size_limit_exceeded"
                    children.append(child_summary)
                    continue
                if context.collector.at_table_capacity:
                    # Keep the full entry manifest, but do not decompress and
                    # parse another child once no result table can be kept.
                    # Its exact table count remains intentionally unknown.
                    context.collector.note_unparsed_source_due_to_table_limit()
                    child_summary["status"] = "not_retained"
                    child_summary["reason"] = "table_limit_reached_before_parse"
                    child_summary["truncated"] = True
                    children.append(child_summary)
                    continue
                try:
                    child_content = archive.read(entry)
                except (OSError, RuntimeError, zipfile.BadZipFile):
                    child_summary["reason"] = "child_read_failed"
                    children.append(child_summary)
                    continue

                content_descriptor = detect_format(child_content, entry.filename)
                if content_descriptor.format == "zip":
                    child_summary["reason"] = "nested_archive_requires_explicit_review"
                    children.append(child_summary)
                    continue

                # A child can contain multiple sheets/tables.  Give its parser
                # only the remaining shared capacity so it does not decode
                # tables which the parent batch is unable to retain.
                child_limits = replace(
                    context.limits,
                    max_tables=context.limits.max_tables - len(context.collector.tables),
                )
                child_result = ingest_file(
                    child_content,
                    entry.filename,
                    limits=child_limits,
                    registry=context.registry,
                    # Archive uploads cannot silently invoke external OCR or a
                    # legacy conversion worker for every image/document.
                    enable_ocr=False,
                    enable_legacy_conversion=False,
                )
                child_tables = child_result.get("tables") or {}
                child_schemas = {
                    schema.get("name"): schema
                    for schema in (child_result.get("manifest") or {}).get("tables") or []
                }
                child_manifest = child_result.get("manifest") or {}
                child_table_limit = child_manifest.get("table_limit") or {}
                child_dropped_tables = max(
                    0, int(child_table_limit.get("dropped_table_count") or 0)
                )
                if child_dropped_tables:
                    # The child parser knows its omitted table count, while the
                    # parent maintains the single, batch-level warning and
                    # truncation contract.
                    context.collector.note_tables_not_retained(child_dropped_tables)
                if child_manifest.get("truncated") and not child_table_limit.get("truncated"):
                    context.collector.note_external_truncation()
                child_summary["status"] = child_result.get("status")
                child_summary["warnings"] = child_result.get("warnings") or []
                child_summary["errors"] = child_result.get("errors") or []
                child_summary["truncated"] = bool(child_manifest.get("truncated"))

                outer_dropped_before = context.collector.dropped_table_count
                for table_name, rows in child_tables.items():
                    schema = child_schemas.get(table_name) or {}
                    columns = schema.get("columns") or list((rows[0] if rows else {}).keys())
                    safe_path_name = re.sub(r"[^A-Za-z0-9._-]+", "_", entry.filename).strip("_")
                    collected_name = context.collector.add_table(
                        f"{safe_path_name}__{table_name}",
                        columns,
                        rows,
                        source_table=f"{entry.filename}:{schema.get('source_table') or table_name}",
                        source_metadata={
                            "archive_path": entry.filename,
                            # Keep the canonical catalog/allowlist key next to
                            # the display path. The evidence API uses this
                            # exact normalized value to select a ZIP child
                            # before parsing unrelated archive members.
                            "normalized_archive_path": normalized_path,
                            "archive_crc": f"{entry.CRC:08x}",
                            "child_format": descriptor.format,
                            "child_source_metadata": schema.get("source_metadata") or {},
                        },
                    )
                    if collected_name:
                        child_summary["tables"].append(collected_name)

                outer_dropped_tables = (
                    context.collector.dropped_table_count - outer_dropped_before
                )
                total_dropped_tables = child_dropped_tables + outer_dropped_tables
                if total_dropped_tables:
                    child_summary["dropped_table_count"] = total_dropped_tables
                    child_summary["reason"] = "some_tables_not_retained_due_to_table_limit"
                    if child_summary["tables"]:
                        child_summary["status"] = "partial"
                    else:
                        child_summary["status"] = "not_retained"

                if not child_tables and not child_summary["errors"]:
                    child_summary["reason"] = "no_structured_tables"
                children.append(child_summary)

        return ParseOutput({
            "batch_manifest": {
                "entry_count": len(file_entries),
                "total_compressed_bytes": sum(entry.compress_size for entry in file_entries),
                "total_uncompressed_bytes": total_uncompressed,
                "entries": manifest_entries,
                "extraction_performed": bool(context.collector.tables),
                "selection": {
                    "archive_entry_allowlist_applied": (
                        context.archive_entry_allowlist is not None
                    ),
                    "selected_normalized_paths": (
                        sorted(context.archive_entry_allowlist)
                        if context.archive_entry_allowlist is not None else None
                    ),
                    "selected_entry_count": (
                        len(context.archive_entry_allowlist)
                        if context.archive_entry_allowlist is not None else len(file_entries)
                    ),
                    "not_selected_entry_count": (
                        len(file_entries) - len(context.archive_entry_allowlist)
                        if context.archive_entry_allowlist is not None else 0
                    ),
                },
                "truncated": (
                    context.collector.table_limit_truncated
                    or bool(context.collector.external_truncated_source_count)
                ),
                "table_limit": context.collector.table_limit_manifest(),
                "children": children,
            }
        })


def _rows_from_dataframe(df: Any) -> Iterator[List[Any]]:
    """Yield data-frame values without exposing pandas/NumPy objects downstream."""

    for row in df.itertuples(index=False, name=None):
        yield list(row)


class OpenXmlSpreadsheetAdapter(IngestionAdapter):
    formats = ("xlsx", "xlsm")

    def capability(self, selected_format: str) -> FormatCapability:
        available = _module_available("openpyxl")
        return FormatCapability(
            format=selected_format,
            family="spreadsheet",
            extensions=(".xlsx", ".xlsm"),
            parse_mode="native_read_only" if available else "optional_dependency",
            supports_tables=True,
            dependency=("openpyxl",),
            available=available,
            notes="Workbooks are opened read-only with formula results only; VBA/macros and external links are never executed.",
        )

    def parse(self, content: bytes, context: ParseContext) -> ParseOutput:
        _safe_zip_entries(content, context.limits)
        if not _module_available("openpyxl"):
            raise IngestionFailure(
                "optional_dependency_missing",
                "XLSX/XLSM parsing requires the openpyxl package.",
                "Install the spreadsheet intake dependency in the worker image.",
            )
        try:
            from openpyxl import load_workbook
            workbook = load_workbook(
                io.BytesIO(content),
                read_only=True,
                data_only=True,
                keep_links=False,
            )
        except Exception as exc:
            raise IngestionFailure(
                "workbook_parse_failed",
                "The XLSX/XLSM workbook could not be opened safely.",
                "Verify the workbook is not corrupted or password-protected.",
            ) from exc

        try:
            worksheets = workbook.worksheets
            for index, worksheet in enumerate(worksheets):
                if context.collector.at_table_capacity:
                    context.collector.note_tables_not_retained(len(worksheets) - index)
                    break
                header, rows = _first_nonempty_row(iter(worksheet.iter_rows(values_only=True)))
                context.collector.add_table(
                    worksheet.title,
                    header,
                    rows,
                    source_table=worksheet.title,
                    source_metadata={"sheet_index": index, "sheet_state": worksheet.sheet_state},
                )
        finally:
            workbook.close()
        return ParseOutput({"workbook_kind": context.source.format, "formula_mode": "cached_values_only"})


class XlsSpreadsheetAdapter(IngestionAdapter):
    formats = ("xls",)

    def capability(self, selected_format: str) -> FormatCapability:
        available = _module_available("xlrd")
        return FormatCapability(
            format="xls",
            family="spreadsheet",
            extensions=(".xls",),
            parse_mode="native_on_demand" if available else "optional_dependency",
            supports_tables=True,
            dependency=("xlrd",),
            available=available,
            notes="Legacy BIFF workbooks are read on demand; macros are not executed.",
        )

    def parse(self, content: bytes, context: ParseContext) -> ParseOutput:
        if not _module_available("xlrd"):
            raise IngestionFailure(
                "optional_dependency_missing",
                "Legacy XLS parsing requires the xlrd package.",
                "Install the spreadsheet intake dependency in the worker image.",
            )
        try:
            import xlrd
            workbook = xlrd.open_workbook(file_contents=content, on_demand=True)
        except Exception as exc:
            raise IngestionFailure(
                "workbook_parse_failed",
                "The XLS workbook could not be opened safely.",
                "Verify the workbook is a valid, non-password-protected XLS file.",
            ) from exc

        def sheet_rows(sheet: Any) -> Iterator[List[Any]]:
            for row_index in range(sheet.nrows):
                values: List[Any] = []
                for column_index in range(sheet.ncols):
                    cell = sheet.cell(row_index, column_index)
                    value = cell.value
                    if cell.ctype == xlrd.XL_CELL_DATE:
                        try:
                            value = xlrd.xldate.xldate_as_datetime(value, workbook.datemode)
                        except (ValueError, OverflowError):
                            pass
                    values.append(value)
                yield values

        try:
            sheet_names = workbook.sheet_names()
            for index, sheet_name in enumerate(sheet_names):
                if context.collector.at_table_capacity:
                    context.collector.note_tables_not_retained(len(sheet_names) - index)
                    break
                sheet = workbook.sheet_by_name(sheet_name)
                header, rows = _first_nonempty_row(sheet_rows(sheet))
                context.collector.add_table(
                    sheet_name,
                    header,
                    rows,
                    source_table=sheet_name,
                    source_metadata={"sheet_index": index},
                )
        finally:
            workbook.release_resources()
        return ParseOutput({"workbook_kind": "xls"})


class OptionalSpreadsheetAdapter(IngestionAdapter):
    formats = ("ods", "xlsb")

    _dependencies = {"ods": ("odf", "odfpy"), "xlsb": ("pyxlsb", "pyxlsb")}

    def capability(self, selected_format: str) -> FormatCapability:
        module, package = self._dependencies[selected_format]
        available = _module_available(module)
        return FormatCapability(
            format=selected_format,
            family="spreadsheet",
            extensions=(f".{selected_format}",),
            parse_mode="optional_dependency" if not available else "native_bounded",
            supports_tables=True,
            dependency=(package,),
            available=available,
            notes="Parsing is enabled only when its local reader dependency is installed; formulas/macros are not executed.",
        )

    def parse(self, content: bytes, context: ParseContext) -> ParseOutput:
        _safe_zip_entries(content, context.limits)
        module, package = self._dependencies[context.source.format]
        if not _module_available(module):
            raise IngestionFailure(
                "optional_dependency_missing",
                f"{context.source.format.upper()} parsing needs the optional '{package}' package.",
                f"Install '{package}' in the approved ingestion-worker image, then retry.",
            )
        # pandas delegates to odfpy/pyxlsb.  Limit every sheet independently
        # before records enter the common table collector.
        try:
            import pandas as pd
            engine = "odf" if context.source.format == "ods" else "pyxlsb"
            with pd.ExcelFile(io.BytesIO(content), engine=engine) as excel_file:
                for index, sheet_name in enumerate(excel_file.sheet_names):
                    if context.collector.at_table_capacity:
                        context.collector.note_tables_not_retained(
                            len(excel_file.sheet_names) - index
                        )
                        break
                    dataframe = pd.read_excel(
                        excel_file,
                        sheet_name=sheet_name,
                        nrows=context.limits.max_rows_per_table + 1,
                    )
                    context.collector.add_table(
                        sheet_name,
                        list(dataframe.columns),
                        _rows_from_dataframe(dataframe),
                        source_table=sheet_name,
                        source_metadata={"sheet_index": index},
                    )
        except IngestionFailure:
            raise
        except Exception as exc:
            raise IngestionFailure(
                "workbook_parse_failed",
                f"The {context.source.format.upper()} workbook could not be parsed.",
                "Verify the workbook is not password-protected or corrupted.",
            ) from exc
        return ParseOutput({"workbook_kind": context.source.format})


class NumbersSpreadsheetAdapter(IngestionAdapter):
    formats = ("numbers",)

    def capability(self, selected_format: str) -> FormatCapability:
        return FormatCapability(
            format="numbers",
            family="spreadsheet",
            extensions=(".numbers",),
            parse_mode="conversion_required",
            supports_tables=False,
            dependency=("approved Numbers-to-XLSX/CSV conversion worker",),
            available=False,
            notes="Apple Numbers packages are detected and ZIP-validated but are not parsed by a generic server worker.",
        )

    def parse(self, content: bytes, context: ParseContext) -> ParseOutput:
        _safe_zip_entries(content, context.limits)
        raise IngestionFailure(
            "numbers_conversion_required",
            "Apple Numbers files require conversion to XLSX or CSV before correlation ingestion.",
            "Export the file as XLSX/CSV in Numbers or route it through an approved conversion worker.",
        )


class ParquetAdapter(IngestionAdapter):
    formats = ("parquet",)

    def capability(self, selected_format: str) -> FormatCapability:
        available = _module_available("pyarrow")
        return FormatCapability(
            format="parquet",
            family="columnar",
            extensions=(".parquet", ".pq"),
            parse_mode="optional_dependency" if not available else "native_row_group_streaming",
            supports_tables=True,
            dependency=("pyarrow",),
            available=available,
            notes="Parquet is read a record batch at a time when pyarrow is installed.",
        )

    def parse(self, content: bytes, context: ParseContext) -> ParseOutput:
        if not _module_available("pyarrow"):
            raise IngestionFailure(
                "optional_dependency_missing",
                "Parquet ingestion needs the optional 'pyarrow' package.",
                "Install pyarrow in the approved ingestion-worker image, then retry.",
            )
        try:
            import pyarrow.parquet as pq
            parquet_file = pq.ParquetFile(io.BytesIO(content))
            schema_names = list(parquet_file.schema_arrow.names)
            if len(schema_names) > context.limits.max_columns:
                raise IngestionFailure(
                    "column_limit_reached",
                    f"Parquet has {len(schema_names)} columns, above the {context.limits.max_columns} safety limit.",
                    "Select a smaller column set in a batch job or increase the approved limit.",
                )

            def record_batches() -> Iterator[Mapping[str, Any]]:
                for batch in parquet_file.iter_batches(batch_size=min(10_000, context.limits.max_rows_per_table)):
                    for record in batch.to_pylist():
                        yield record

            context.collector.add_table(
                "parquet",
                schema_names,
                record_batches(),
                source_table="parquet",
                source_metadata={
                    "row_groups": parquet_file.num_row_groups,
                    "declared_rows": parquet_file.metadata.num_rows if parquet_file.metadata else None,
                },
            )
            return ParseOutput({"columnar_encoding": "parquet", "row_groups": parquet_file.num_row_groups})
        except IngestionFailure:
            raise
        except Exception as exc:
            raise IngestionFailure(
                "parquet_parse_failed",
                "The file could not be read as a valid Parquet dataset.",
                "Verify that the upload is a complete Parquet file.",
            ) from exc


class ArrowAdapter(IngestionAdapter):
    formats = ("arrow",)

    def capability(self, selected_format: str) -> FormatCapability:
        available = _module_available("pyarrow")
        return FormatCapability(
            format="arrow",
            family="columnar",
            extensions=(".arrow", ".feather", ".ipc"),
            parse_mode="optional_dependency" if not available else "native_record_batch_streaming",
            supports_tables=True,
            dependency=("pyarrow",),
            available=available,
            notes="Apache Arrow IPC/Feather requires pyarrow and is read one record batch at a time.",
        )

    def parse(self, content: bytes, context: ParseContext) -> ParseOutput:
        if not _module_available("pyarrow"):
            raise IngestionFailure(
                "optional_dependency_missing",
                "Arrow/Feather ingestion needs the optional 'pyarrow' package.",
                "Install pyarrow in the approved ingestion-worker image, then retry.",
            )
        try:
            import pyarrow as pa

            reader: Any
            stream_mode = False
            input_buffer = pa.BufferReader(content)
            try:
                reader = pa.ipc.open_file(input_buffer)
            except pa.ArrowInvalid:
                # Not a failure — a format probe. Arrow IPC has two container formats and
                # only trying one tells them apart, so the raised `ArrowInvalid` IS the
                # answer "this is a stream". Caught narrowly on purpose: a broad handler
                # here would take a genuine corrupt-file error and retry it as a stream,
                # reporting whatever the second parse said about the wrong format.
                stream_mode = True
                reader = pa.ipc.open_stream(pa.BufferReader(content))
            schema_names = list(reader.schema.names)
            if len(schema_names) > context.limits.max_columns:
                raise IngestionFailure(
                    "column_limit_reached",
                    f"Arrow data has {len(schema_names)} columns, above the {context.limits.max_columns} safety limit.",
                    "Select a smaller column set in a batch job or increase the approved limit.",
                )

            def record_batches() -> Iterator[Mapping[str, Any]]:
                if stream_mode:
                    for batch in reader:
                        for record in batch.to_pylist():
                            yield record
                else:
                    for index in range(reader.num_record_batches):
                        batch = reader.get_batch(index)
                        for record in batch.to_pylist():
                            yield record

            context.collector.add_table(
                "arrow",
                schema_names,
                record_batches(),
                source_table="arrow",
                source_metadata={"stream_mode": stream_mode},
            )
            return ParseOutput({"columnar_encoding": "arrow", "stream_mode": stream_mode})
        except IngestionFailure:
            raise
        except Exception as exc:
            raise IngestionFailure(
                "arrow_parse_failed",
                "The file could not be read as Arrow IPC/Feather data.",
                "Verify that the upload is a complete Arrow IPC or Feather file.",
            ) from exc


def _add_document_tables(
    context: ParseContext,
    tables: Sequence[Mapping[str, Any]],
    *,
    location_key: str,
) -> None:
    """Turn parser table structures into normal evidence-ready tables."""

    nonempty_tables = [
        (index, table)
        for index, table in enumerate(tables)
        if table.get("rows")
    ]
    for retained_index, (index, table) in enumerate(nonempty_tables):
        if context.collector.at_table_capacity:
            context.collector.note_tables_not_retained(len(nonempty_tables) - retained_index)
            break
        raw_rows = table.get("rows") or []
        header = list(raw_rows[0]) if raw_rows else []
        data_rows = raw_rows[1:] if len(raw_rows) > 1 else []
        source_location = table.get(location_key, index + 1)
        context.collector.add_table(
            f"table_{index + 1}",
            header,
            data_rows,
            source_table=f"table_{index + 1}",
            source_metadata={location_key: source_location},
        )


class DocumentTableAdapter(IngestionAdapter):
    """Bridge existing deterministic PDF/DOCX extractors to the table contract."""

    formats = ("pdf", "docx")

    def capability(self, selected_format: str) -> FormatCapability:
        if selected_format == "pdf":
            available = _module_available("pdfplumber")
            dependency = ("pdfplumber", "pypdf")
        else:
            available = _module_available("docx")
            dependency = ("python-docx",)
        return FormatCapability(
            format=selected_format,
            family="document",
            extensions=(f".{selected_format}",),
            parse_mode="existing_structural_parser" if available else "optional_dependency",
            supports_tables=True,
            dependency=dependency,
            available=available,
            notes="Extracts embedded tables deterministically; OCR is a separate opt-in adapter for scanned content.",
        )

    def parse(self, content: bytes, context: ParseContext) -> ParseOutput:
        try:
            if context.source.format == "pdf":
                from app.services.pdf_parser import parse_pdf_structure
                structure = parse_pdf_structure(
                    content,
                    context.filename,
                    max_pages=context.limits.max_document_pages,
                )
                _add_document_tables(context, structure.get("tables") or [], location_key="page_num")
            else:
                # DOCX is a ZIP package, so apply the same archive-bomb/path
                # checks before handing it to python-docx.
                _safe_zip_entries(content, context.limits)
                from app.services.docx_parser import parse_docx_structure
                structure = parse_docx_structure(content, context.filename)
                _add_document_tables(context, structure.get("tables") or [], location_key="section_id")
        except Exception as exc:
            raise IngestionFailure(
                "document_parse_failed",
                f"The {context.source.format.upper()} document could not be parsed.",
                "Verify that the document is not corrupted or password-protected.",
            ) from exc
        return ParseOutput({
            "document_structure": {
                "subtype": structure.get("subtype"),
                "shared_keys": structure.get("shared_keys", []),
                "truncated": structure.get("truncated", False),
            }
        })


class LegacyDocAdapter(IngestionAdapter):
    formats = ("doc",)

    def capability(self, selected_format: str) -> FormatCapability:
        return FormatCapability(
            format="doc",
            family="document",
            extensions=(".doc",),
            parse_mode="approved_conversion_worker",
            supports_tables=True,
            dependency=("LegacyDocConverter",),
            available=False,
            notes="Legacy Word files are not parsed in-process; a registered converter must be explicitly enabled.",
        )

    def parse(self, content: bytes, context: ParseContext) -> ParseOutput:
        converter = context.registry.legacy_doc_converter
        if not context.enable_legacy_conversion or converter is None:
            raise IngestionFailure(
                "legacy_doc_conversion_required",
                "Legacy .doc files need an approved conversion worker before they can be correlated.",
                "Convert to DOCX/PDF or enable a registered LegacyDocConverter in a worker job.",
            )
        try:
            converted = dict(converter.convert(content, context.filename, limits=context.limits))
        except Exception as exc:
            raise IngestionFailure(
                "legacy_doc_conversion_failed",
                "The approved legacy-document converter could not process this file.",
                "Check the conversion worker logs or convert the file to DOCX manually.",
                retryable=True,
            ) from exc
        _merge_external_adapter_result(converted, context, source="legacy_doc_converter")
        return ParseOutput({"conversion_adapter": getattr(converter, "name", converter.__class__.__name__)})


class ImageOcrAdapter(IngestionAdapter):
    formats = ("image",)

    def capability(self, selected_format: str) -> FormatCapability:
        return FormatCapability(
            format="image",
            family="image",
            extensions=(".png", ".jpg", ".jpeg", ".tif", ".tiff", ".bmp", ".webp"),
            parse_mode="opt_in_ocr_adapter",
            supports_tables=True,
            dependency=("OCRAdapter",),
            available=False,
            notes="OCR is never sent to an external provider unless an approved adapter is registered and enabled.",
        )

    def parse(self, content: bytes, context: ParseContext) -> ParseOutput:
        adapter = context.registry.ocr_adapter
        if not context.enable_ocr or adapter is None:
            context.warnings.append(IngestionIssue(
                "ocr_not_run",
                "Image content was detected, but OCR is not enabled for this ingestion job.",
                "Register an approved OCRAdapter and call ingest_file(..., enable_ocr=True) to extract text/tables.",
            ))
            return ParseOutput({"ocr": {"status": "not_requested"}})
        try:
            extracted = dict(adapter.extract(content, context.filename, limits=context.limits))
        except Exception as exc:
            raise IngestionFailure(
                "ocr_failed",
                "The approved OCR adapter could not process this image.",
                "Check the OCR worker/service and retry the job.",
                retryable=True,
            ) from exc
        _merge_external_adapter_result(extracted, context, source="ocr")
        return ParseOutput({"ocr": {"status": "completed", "adapter": getattr(adapter, "name", adapter.__class__.__name__)}})


class UnknownFormatAdapter(IngestionAdapter):
    formats = ("unknown",)

    def capability(self, selected_format: str) -> FormatCapability:
        return FormatCapability(
            format="unknown",
            family="unknown",
            extensions=(),
            parse_mode="not_supported",
            supports_tables=False,
            available=False,
            notes="No parser is selected until the file is exported to a supported interchange format.",
        )

    def parse(self, content: bytes, context: ParseContext) -> ParseOutput:
        raise IngestionFailure(
            "unsupported_format",
            "The file format could not be identified safely.",
            "Upload CSV, Excel, JSON/JSONL, Parquet, Arrow, XML, PDF/DOCX, or a ZIP batch manifest.",
        )


def _merge_external_adapter_result(
    result: Mapping[str, Any], context: ParseContext, *, source: str
) -> None:
    """Merge a deliberately registered OCR/conversion result into tables.

    Adapters only get to contribute primitive table records and warnings.  They
    cannot override the source hash, format detection, limits, or capability
    metadata generated by this service.
    """

    raw_tables = result.get("tables") or {}
    if not isinstance(raw_tables, Mapping):
        raise IngestionFailure(
            "adapter_result_invalid",
            f"The {source} adapter returned an invalid tables payload.",
        )
    table_items = list(raw_tables.items())
    for index, (table_name, records) in enumerate(table_items):
        if context.collector.at_table_capacity:
            context.collector.note_tables_not_retained(len(table_items) - index)
            break
        if not isinstance(records, Sequence) or isinstance(records, (str, bytes, bytearray)):
            raise IngestionFailure(
                "adapter_result_invalid",
                f"The {source} adapter returned a non-list table '{table_name}'.",
            )
        mapping_rows = [dict(row) if isinstance(row, Mapping) else {"value": row} for row in records]
        columns, _ = _columns_from_mapping_rows(mapping_rows, context.limits.max_columns)
        context.collector.add_table(
            table_name,
            columns,
            mapping_rows,
            source_table=str(table_name),
            source_metadata={"adapter": source},
        )
    for warning in result.get("warnings") or []:
        if isinstance(warning, Mapping):
            context.warnings.append(IngestionIssue(
                str(warning.get("code") or "adapter_warning"),
                str(warning.get("message") or "The adapter reported a warning."),
                str(warning["remediation"]) if warning.get("remediation") else None,
            ))


# ---------------------------------------------------------------------------
# External operational-source contracts (intentionally planning-only here)
# ---------------------------------------------------------------------------


@dataclass(frozen=True)
class ConnectorCapability:
    """Metadata for a connector type that an approved worker can implement."""

    key: str
    category: str
    display_name: str
    required_config_keys: Tuple[str, ...]
    supports_incremental: bool
    supports_streaming: bool
    notes: str = ""

    def as_dict(self) -> Dict[str, Any]:
        return {
            "key": self.key,
            "category": self.category,
            "display_name": self.display_name,
            "required_config_keys": list(self.required_config_keys),
            "supports_incremental": self.supports_incremental,
            "supports_streaming": self.supports_streaming,
            "connection_attempted": False,
            "notes": self.notes,
        }


class OperationalSourceConnector(ABC):
    """Base interface for database/ERP/historian/stream source adapters.

    This ingestion module only produces a *sanitized extraction plan*.  A
    credential-aware worker should subclass this contract and implement its own
    explicitly authorized extraction path.  That separation prevents an upload
    parsing endpoint from unexpectedly dialing out to customer infrastructure.
    """

    @property
    @abstractmethod
    def capability(self) -> ConnectorCapability:
        ...

    @abstractmethod
    def build_plan(
        self,
        configuration: Optional[Mapping[str, Any]] = None,
        *,
        entities: Optional[Sequence[str]] = None,
        cursor: Optional[str] = None,
    ) -> Dict[str, Any]:
        ...


class DatabaseConnector(OperationalSourceConnector, ABC):
    """Contract marker for relational databases."""


class WarehouseConnector(OperationalSourceConnector, ABC):
    """Contract marker for analytical warehouses/lakes."""


class ERPSourceConnector(OperationalSourceConnector, ABC):
    """Contract marker for ERP sources and ERP export APIs."""


class HistorianConnector(OperationalSourceConnector, ABC):
    """Contract marker for industrial historians/time-series systems."""


class EventStreamConnector(OperationalSourceConnector, ABC):
    """Contract marker for event streams and message brokers."""


class _PlanOnlyConnector(OperationalSourceConnector):
    """Safe catalog implementation that never retains configuration values."""

    def __init__(self, capability: ConnectorCapability) -> None:
        self._capability = capability

    @property
    def capability(self) -> ConnectorCapability:
        return self._capability

    def build_plan(
        self,
        configuration: Optional[Mapping[str, Any]] = None,
        *,
        entities: Optional[Sequence[str]] = None,
        cursor: Optional[str] = None,
    ) -> Dict[str, Any]:
        config = dict(configuration or {})
        present_keys = sorted(str(key) for key in config)
        missing = [key for key in self.capability.required_config_keys if key not in config]
        # The plan deliberately reports only key names.  URLs, passwords,
        # tokens, connection strings and other values must not travel through
        # a capability/preview response.
        return {
            "connector": self.capability.as_dict(),
            "status": "requires_approved_worker",
            "provided_configuration_keys": present_keys,
            "missing_required_configuration_keys": missing,
            "requested_entities": [str(entity) for entity in (entities or [])][:100],
            "cursor_supplied": bool(cursor),
            "connection_attempted": False,
        }


_CONNECTOR_CAPABILITIES: Tuple[ConnectorCapability, ...] = (
    ConnectorCapability("postgres", "database", "PostgreSQL", ("host", "database", "credential_ref"), True, False),
    ConnectorCapability("mysql", "database", "MySQL", ("host", "database", "credential_ref"), True, False),
    ConnectorCapability("mssql", "database", "Microsoft SQL Server", ("host", "database", "credential_ref"), True, False),
    ConnectorCapability("oracle_database", "database", "Oracle Database", ("host", "service", "credential_ref"), True, False),
    ConnectorCapability("snowflake", "warehouse", "Snowflake", ("account", "warehouse", "credential_ref"), True, False),
    ConnectorCapability("bigquery", "warehouse", "BigQuery", ("project", "dataset", "credential_ref"), True, False),
    ConnectorCapability("databricks", "warehouse", "Databricks", ("workspace", "warehouse", "credential_ref"), True, False),
    ConnectorCapability("redshift", "warehouse", "Amazon Redshift", ("host", "database", "credential_ref"), True, False),
    ConnectorCapability("sap", "erp", "SAP", ("base_url", "credential_ref"), True, True),
    ConnectorCapability("oracle_erp", "erp", "Oracle ERP", ("base_url", "credential_ref"), True, True),
    ConnectorCapability("dynamics", "erp", "Microsoft Dynamics", ("base_url", "credential_ref"), True, True),
    ConnectorCapability("netsuite", "erp", "NetSuite", ("account", "credential_ref"), True, True),
    ConnectorCapability("odoo", "erp", "Odoo", ("base_url", "credential_ref"), True, True),
    ConnectorCapability("osi_pi", "historian", "OSIsoft PI / AVEVA PI", ("host", "credential_ref"), True, True),
    ConnectorCapability("influxdb", "historian", "InfluxDB", ("url", "bucket", "credential_ref"), True, True),
    ConnectorCapability("timescaledb", "historian", "TimescaleDB", ("host", "database", "credential_ref"), True, True),
    ConnectorCapability("opcua", "historian", "OPC UA", ("endpoint", "credential_ref"), True, True),
    ConnectorCapability("kafka", "event_stream", "Apache Kafka", ("brokers", "topic", "credential_ref"), True, True),
    ConnectorCapability("mqtt", "event_stream", "MQTT", ("broker", "topic", "credential_ref"), True, True),
    ConnectorCapability("webhook", "event_stream", "Webhook", ("endpoint", "credential_ref"), False, True),
    ConnectorCapability("rest", "event_stream", "REST API", ("base_url", "credential_ref"), True, False),
)


class ConnectorCatalog:
    """Catalog and plan factory that exposes no connection side effects."""

    def __init__(self, capabilities: Sequence[ConnectorCapability] = _CONNECTOR_CAPABILITIES) -> None:
        self._connectors = {capability.key: _PlanOnlyConnector(capability) for capability in capabilities}

    def capabilities(self) -> List[Dict[str, Any]]:
        return [self._connectors[key].capability.as_dict() for key in sorted(self._connectors)]

    def plan(
        self,
        key: str,
        configuration: Optional[Mapping[str, Any]] = None,
        *,
        entities: Optional[Sequence[str]] = None,
        cursor: Optional[str] = None,
    ) -> Dict[str, Any]:
        connector = self._connectors.get(str(key).strip().lower())
        if connector is None:
            return {
                "status": "unsupported_connector",
                "error": _issue(
                    "unsupported_connector",
                    f"'{key}' is not in the operational source connector catalog.",
                    "Choose a listed connector type or register an approved worker adapter.",
                ),
                "available_connectors": sorted(self._connectors),
                "connection_attempted": False,
            }
        return connector.build_plan(configuration, entities=entities, cursor=cursor)


DEFAULT_CONNECTOR_CATALOG = ConnectorCatalog()


def connector_capabilities() -> List[Dict[str, Any]]:
    """Return database/warehouse/ERP/historian/event-stream capability metadata."""

    return DEFAULT_CONNECTOR_CATALOG.capabilities()


def plan_connector_ingestion(
    connector: str,
    configuration: Optional[Mapping[str, Any]] = None,
    *,
    entities: Optional[Sequence[str]] = None,
    cursor: Optional[str] = None,
) -> Dict[str, Any]:
    """Return a sanitized, non-networked plan for an external source."""

    return DEFAULT_CONNECTOR_CATALOG.plan(connector, configuration, entities=entities, cursor=cursor)


# ---------------------------------------------------------------------------
# Registry construction and public file-ingestion API
# ---------------------------------------------------------------------------


def create_default_registry() -> IngestionAdapterRegistry:
    """Create an isolated registry suitable for tests or a custom worker."""

    registry = IngestionAdapterRegistry()
    for adapter in (
        DelimitedAdapter(),
        JsonAdapter(),
        XmlAdapter(),
        OpenXmlSpreadsheetAdapter(),
        XlsSpreadsheetAdapter(),
        OptionalSpreadsheetAdapter(),
        NumbersSpreadsheetAdapter(),
        ParquetAdapter(),
        ArrowAdapter(),
        ZipBatchAdapter(),
        DocumentTableAdapter(),
        LegacyDocAdapter(),
        ImageOcrAdapter(),
        UnknownFormatAdapter(),
    ):
        registry.register(adapter)
    return registry


DEFAULT_INGESTION_REGISTRY = create_default_registry()


def capability_manifest(
    registry: Optional[IngestionAdapterRegistry] = None,
) -> Dict[str, Any]:
    """Describe all locally configured format and external-source capabilities."""

    active_registry = registry or DEFAULT_INGESTION_REGISTRY
    return {
        "formats": active_registry.capabilities(),
        "external_connectors": connector_capabilities(),
        "ocr": {
            "configured": active_registry.ocr_adapter is not None,
            "adapter": getattr(active_registry.ocr_adapter, "name", None),
            "automatic_execution": False,
        },
        "legacy_doc_conversion": {
            "configured": active_registry.legacy_doc_converter is not None,
            "adapter": getattr(active_registry.legacy_doc_converter, "name", None),
            "automatic_execution": False,
        },
    }


def _base_result(
    *,
    content: bytes,
    filename: str,
    descriptor: FormatDescriptor,
    limits: Optional[IngestionLimits],
    capability: Optional[FormatCapability],
    registry: IngestionAdapterRegistry,
) -> Dict[str, Any]:
    return {
        "status": "rejected",
        "tables": {},
        "manifest": {
            "source": {
                "filename": filename or "upload",
                "size_bytes": len(content),
                "sha256": hashlib.sha256(content).hexdigest(),
                "format": descriptor.as_dict(),
            },
            "limits": limits.as_dict() if limits else None,
            "table_count": 0,
            "row_count": 0,
            "tables": [],
            "truncated": False,
        },
        "capabilities": {
            "selected_format": capability.as_dict() if capability else None,
            "ocr": {
                "configured": registry.ocr_adapter is not None,
                "adapter": getattr(registry.ocr_adapter, "name", None),
                "executed": False,
            },
            "legacy_doc_conversion": {
                "configured": registry.legacy_doc_converter is not None,
                "adapter": getattr(registry.legacy_doc_converter, "name", None),
                "executed": False,
            },
            # This compact category list lets an API advertise expansion paths
            # without returning every connector definition with every upload.
            "external_source_categories": ["database", "warehouse", "erp", "historian", "event_stream"],
        },
        "warnings": [],
        "errors": [],
    }


def ingest_file(
    content: bytes,
    filename: str,
    limits: LimitsInput = None,
    content_type: Optional[str] = None,
    *,
    registry: Optional[IngestionAdapterRegistry] = None,
    enable_ocr: bool = False,
    enable_legacy_conversion: bool = False,
    archive_entry_allowlist: ArchiveEntryAllowlistInput = None,
) -> Dict[str, Any]:
    """Safely ingest one local file into a common table-shaped result.

    Parameters are intentionally dependency-light: callers only need bytes and
    a filename.  ``limits`` accepts :class:`IngestionLimits` or a mapping of
    its documented fields.  The function never raises for a bad upload or a
    missing optional parser; inspect ``status`` and ``errors`` instead.

    Result contract (stable keys):

    ``tables``
        ``{table_name: list[dict]}``, suitable for typed/evidence
        normalization.  Values are JSON-safe and bounded.
    ``manifest``
        Source hash, detector result, parser extras, table schemas and row
        truncation/lineage metadata.
    ``capabilities`` / ``warnings`` / ``errors``
        Explicit capability state and non-fatal/fatal diagnostics.

    ``archive_entry_allowlist``
        Optional exact normalized member paths for a top-level ZIP batch.  It
        is rejected for non-ZIP inputs and never applies recursively.
    """

    active_registry = registry or DEFAULT_INGESTION_REGISTRY
    if not isinstance(content, (bytes, bytearray, memoryview)):
        # Do not call bytes() on arbitrary objects; it could invoke a surprising
        # user implementation or allocate an unbounded stream in an API path.
        descriptor = FormatDescriptor("unknown", "unknown")
        result = _base_result(
            content=b"", filename=filename, descriptor=descriptor, limits=None,
            capability=None, registry=active_registry,
        )
        result["errors"].append(_issue(
            "invalid_content_type", "ingest_file expects an in-memory bytes payload."
        ))
        return result

    payload = bytes(content)
    descriptor = detect_format(payload, filename, content_type)
    try:
        parsed_limits = _coerce_limits(limits)
    except ValueError as exc:
        result = _base_result(
            content=payload, filename=filename, descriptor=descriptor, limits=None,
            capability=None, registry=active_registry,
        )
        result["errors"].append(_issue("invalid_limits", str(exc)))
        return result

    adapter = active_registry.get(descriptor.format) or active_registry.get("unknown")
    capability = adapter.capability(descriptor.format) if adapter else None
    result = _base_result(
        content=payload,
        filename=filename,
        descriptor=descriptor,
        limits=parsed_limits,
        capability=capability,
        registry=active_registry,
    )
    if archive_entry_allowlist is not None and descriptor.format != "zip":
        result["errors"].append(_issue(
            "archive_entry_allowlist_requires_zip",
            "archive_entry_allowlist can only select children of a ZIP batch.",
            "Remove the selection or upload a ZIP batch.",
        ))
        return result
    try:
        parsed_archive_entry_allowlist = _coerce_archive_entry_allowlist(
            archive_entry_allowlist,
            max_entries=parsed_limits.max_zip_entries,
        )
    except ValueError as exc:
        result["errors"].append(_issue("invalid_archive_entry_allowlist", str(exc)))
        return result
    if len(payload) > parsed_limits.max_file_bytes:
        result["errors"].append(_issue(
            "file_size_limit_exceeded",
            f"The file is {len(payload)} bytes, above the {parsed_limits.max_file_bytes} byte ingestion limit.",
            "Use an asynchronous batch job, split the file, or raise the approved limit.",
        ))
        return result
    if not payload:
        result["errors"].append(_issue("empty_file", "The uploaded file is empty."))
        return result
    if descriptor.mismatch:
        result["warnings"].append(_issue(
            "filename_content_mismatch",
            f"The filename suggests {descriptor.format}, but file bytes resemble {descriptor.magic_format}.",
            "Verify that the extension matches the actual export before relying on the result.",
        ))

    collector = TableCollector(parsed_limits)
    context = ParseContext(
        source=descriptor,
        filename=filename or "upload",
        limits=parsed_limits,
        collector=collector,
        registry=active_registry,
        enable_ocr=enable_ocr,
        enable_legacy_conversion=enable_legacy_conversion,
        archive_entry_allowlist=parsed_archive_entry_allowlist,
    )
    try:
        if adapter is None:  # defensive: default registry always has unknown
            raise IngestionFailure("unsupported_format", "No ingestion adapter is registered for this format.")
        output = adapter.parse(payload, context)
    except IngestionFailure as exc:
        result["errors"].append(exc.issue.as_dict())
        result["warnings"].extend(
            issue.as_dict() for issue in (context.warnings + collector.all_warnings())
        )
        return result
    except Exception:
        # Avoid leaking parser internals to the API response; the caller still
        # receives a deterministic remediation and application logs can retain
        # the actual exception at the integration boundary.
        result["errors"].append(_issue(
            "ingestion_parse_failed",
            "The file could not be parsed by the selected ingestion adapter.",
            "Verify the source file and retry; contact support with the source hash if it persists.",
        ))
        result["warnings"].extend(
            issue.as_dict() for issue in (context.warnings + collector.all_warnings())
        )
        return result

    result["tables"] = dict(collector.tables)
    result["manifest"].update(output.manifest)
    result["manifest"]["tables"] = list(collector.schemas.values())
    result["manifest"]["table_count"] = len(collector.tables)
    result["manifest"]["row_count"] = collector.total_rows
    result["manifest"]["table_limit"] = collector.table_limit_manifest()
    result["manifest"]["truncated"] = (
        any(schema["truncated"] for schema in collector.schemas.values())
        or collector.table_limit_truncated
        or bool(collector.external_truncated_source_count)
    )
    result["warnings"].extend(
        issue.as_dict() for issue in (context.warnings + collector.all_warnings())
    )
    result["capabilities"]["ocr"]["executed"] = bool(enable_ocr and active_registry.ocr_adapter is not None and descriptor.format == "image")
    result["capabilities"]["legacy_doc_conversion"]["executed"] = bool(
        enable_legacy_conversion and active_registry.legacy_doc_converter is not None and descriptor.format == "doc"
    )

    if result["manifest"].get("batch_manifest") and not result["tables"]:
        result["status"] = "manifested"
    elif descriptor.format == "image" and not result["tables"]:
        result["status"] = "awaiting_ocr"
    elif result["manifest"]["truncated"] or result["warnings"]:
        result["status"] = "partial"
    else:
        result["status"] = "parsed"
    return result


__all__ = [
    "ArchiveEntryAllowlistInput",
    "ArrowAdapter",
    "ConnectorCapability",
    "ConnectorCatalog",
    "DatabaseConnector",
    "DEFAULT_CONNECTOR_CATALOG",
    "DEFAULT_INGESTION_REGISTRY",
    "ERPSourceConnector",
    "EventStreamConnector",
    "FormatCapability",
    "FormatDescriptor",
    "HistorianConnector",
    "IngestionAdapter",
    "IngestionAdapterRegistry",
    "IngestionFailure",
    "IngestionIssue",
    "IngestionLimits",
    "LegacyDocConverter",
    "OCRAdapter",
    "OperationalSourceConnector",
    "WarehouseConnector",
    "capability_manifest",
    "connector_capabilities",
    "create_default_registry",
    "detect_format",
    "ingest_file",
    "plan_connector_ingestion",
]
